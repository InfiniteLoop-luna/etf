"""Detect and repair missing raw values in the broad-index ETF workbook."""

from __future__ import annotations

import argparse
import sys
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from main import main as update_one_date  # noqa: E402
from main import load_data_sources  # noqa: E402
from src.data_source_manager import DataSourceManager  # noqa: E402

DATE_ROW = 2
RAW_ROWS = tuple(range(3, 16)) + tuple(range(19, 32))
EXPECTED_RAW_VALUES = len(RAW_ROWS)


def raw_value_count(ws, column: int) -> int:
    return sum(isinstance(ws.cell(row, column).value, (int, float)) for row in RAW_ROWS)


def find_incomplete_dates(workbook: Path, start: date, end: date) -> list[str]:
    wb = openpyxl.load_workbook(workbook, data_only=False, read_only=True)
    ws = wb.active
    incomplete: list[str] = []
    for column in range(3, ws.max_column + 1):
        value = ws.cell(DATE_ROW, column).value
        if isinstance(value, datetime):
            current = value.date()
        elif isinstance(value, date):
            current = value
        elif isinstance(value, str):
            try:
                current = datetime.strptime(value.replace('/', '-'), '%Y-%m-%d').date()
            except ValueError:
                continue
        else:
            continue
        if start <= current <= end and raw_value_count(ws, column) < EXPECTED_RAW_VALUES:
            incomplete.append(current.isoformat())
    wb.close()
    return incomplete


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument('--excel', default='主要ETF基金份额变动情况.xlsx')
    parser.add_argument('--lookback-days', type=int, default=45)
    parser.add_argument('--end-date', help='YYYY-MM-DD; defaults to today')
    args = parser.parse_args()

    end = datetime.strptime(args.end_date, '%Y-%m-%d').date() if args.end_date else date.today()
    start = end - timedelta(days=max(args.lookback_days, 0))
    workbook = Path(args.excel)
    candidates = find_incomplete_dates(workbook, start, end)
    if not candidates:
        print(f'ETF Excel integrity OK: {start}..{end}')
        return 0

    source_manager = DataSourceManager(load_data_sources())
    trading_gaps = [day for day in candidates if source_manager.is_trading_day(day)]
    print(f'ETF Excel trading-day gaps: {trading_gaps}')
    for day in trading_gaps:
        status = update_one_date(day)
        if status != 0:
            print(f'Failed to repair {day}: main.py status={status}', file=sys.stderr)
            return status

    remaining = set(find_incomplete_dates(workbook, start, end))
    unresolved = [day for day in trading_gaps if day in remaining]
    if unresolved:
        print(f'Unresolved ETF Excel trading-day gaps: {unresolved}', file=sys.stderr)
        return 4
    print(f'ETF Excel repaired: {trading_gaps}')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
