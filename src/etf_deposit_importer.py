from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


LABEL_TO_FIELD = {
    "人民币存款余额": "rmb_deposit_balance",
    "外币存款余额": "fx_deposit_balance",
    "本外币存款余额": "total_deposit_balance",
    "住户存款增加额": "household_deposit_increase",
    "非金融企业存款增加额": "corp_deposit_increase",
    "财政性存款增加额": "fiscal_deposit_increase",
    "非银行业金融机构存款增加额": "nonbank_deposit_increase",
    "存款合计增加额": "total_deposit_increase",
    "居民长期贷款增加额": "household_long_loan_increase",
}


def parse_deposit_workbook(path) -> pd.DataFrame:
    workbook_source = Path(path) if isinstance(path, (str, Path)) else path
    wb = load_workbook(workbook_source, data_only=True, read_only=True)
    try:
        if "本外币存款数据" not in wb.sheetnames:
            raise ValueError("导入文件缺少 本外币存款数据 sheet")

        ws = wb["本外币存款数据"]
        rows = []
        current = None

        for row_idx in range(1, ws.max_row + 1):
            month_value = ws.cell(row_idx, 2).value
            label = ws.cell(row_idx, 3).value
            number = ws.cell(row_idx, 4).value

            if month_value is not None:
                if current and len(current) > 1:
                    rows.append(current)
                current = {"month": pd.to_datetime(month_value).strftime("%Y-%m-%d")}

            if current is not None and label in LABEL_TO_FIELD:
                current[LABEL_TO_FIELD[label]] = float(number) if number is not None else None

        if current and len(current) > 1:
            rows.append(current)

        df = pd.DataFrame(rows)
        if df.empty:
            raise ValueError("未解析出任何月份数据")
        return df.sort_values("month").reset_index(drop=True)
    finally:
        wb.close()
