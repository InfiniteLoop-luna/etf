"""Excel管理器 - 动态识别section并更新数据"""

import openpyxl
import logging
from dataclasses import dataclass
from typing import Dict, List, Optional


@dataclass
class Section:
    """Excel中的一个section"""
    name: str           # Section名称（如"总市值：亿元"）
    header_row: int     # 标题行号
    data_start: int     # 数据起始行号
    data_end: int       # 数据结束行号

    @property
    def is_calculated(self) -> bool:
        """判断是否为计算型section"""
        calculated_keywords = ['份额', '变动', '申赎', '比例', '涨跌幅']
        return any(kw in self.name for kw in calculated_keywords)

    @property
    def is_data_section(self) -> bool:
        """判断是否为需要更新的数据section"""
        return '总市值' in self.name or '基金单位市值' in self.name


class DynamicExcelManager:
    """动态Excel管理器，自动识别section结构"""

    # 常量定义
    HEADER_ROW = 1
    DATE_ROW = 2
    CODE_COL = 1
    NAME_COL = 2
    DATA_START_COL = 3

    def __init__(self, file_path: str):
        self.file_path = file_path
        self.logger = logging.getLogger(__name__)
        self.wb = openpyxl.load_workbook(file_path)
        self.ws = self.wb.active
        self.sections = self._detect_sections()

    def _detect_sections(self) -> Dict[str, Section]:
        """动态扫描并识别所有section"""
        sections = {}
        current_section = None

        # 特殊处理第一个section（总市值）
        # 第一行是标题行，第二行是日期行，第三行开始是数据
        first_section_name = self.ws.cell(self.HEADER_ROW, self.DATA_START_COL).value
        if first_section_name and isinstance(first_section_name, str):
            current_section = Section(
                name=first_section_name,
                header_row=self.HEADER_ROW,
                data_start=self.HEADER_ROW + 2,  # 跳过标题行和日期行
                data_end=None
            )

        for row_idx in range(1, self.ws.max_row + 1):
            if self._is_section_header(row_idx):
                # 结束上一个section
                if current_section:
                    current_section.data_end = row_idx - 1
                    sections[current_section.name] = current_section

                # 开始新section
                section_name = self.ws.cell(row_idx, self.NAME_COL).value
                current_section = Section(
                    name=section_name,
                    header_row=row_idx,
                    data_start=row_idx + 1,
                    data_end=None
                )

        # 处理最后一个section
        if current_section:
            current_section.data_end = self.ws.max_row
            sections[current_section.name] = current_section

        self.logger.info(f"检测到 {len(sections)} 个section")
        return sections

    def _is_section_header(self, row: int) -> bool:
        """判断是否为section标题行"""
        code_cell = self.ws.cell(row, self.CODE_COL).value
        name_cell = self.ws.cell(row, self.NAME_COL).value

        # Section header特征：第0列为空，第1列包含关键词
        keywords = ['市值', '份额', '变动', '申赎', '比例', '涨跌幅']
        return (code_cell is None and
                name_cell and
                isinstance(name_cell, str) and
                any(kw in name_cell for kw in keywords))

    def get_etf_codes(self) -> List[str]:
        """从第一个数据section获取所有ETF代码"""
        # 找到第一个数据section
        data_section = next(
            (s for s in self.sections.values() if s.is_data_section),
            None
        )

        if not data_section:
            raise ValueError("未找到数据section")

        codes = []
        for row in range(data_section.data_start, data_section.data_end + 1):
            code = self.ws.cell(row, self.CODE_COL).value
            if code and isinstance(code, str):
                codes.append(code)

        return codes

    def find_or_create_date_column(self, target_date: str) -> int:
        """查找或创建日期列"""
        from datetime import datetime

        # 在DATE_ROW查找日期
        for col in range(self.DATA_START_COL, self.ws.max_column + 1):
            date_val = self.ws.cell(self.DATE_ROW, col).value

            # 处理datetime对象
            if isinstance(date_val, datetime):
                date_str = date_val.strftime('%Y-%m-%d')
                if date_str == target_date:
                    return col
            # 处理字符串
            elif isinstance(date_val, str):
                if date_val == target_date:
                    return col

        # 未找到，在最后添加新列
        new_col = self.ws.max_column + 1
        # 将日期作为datetime对象存储，保持与现有格式一致
        date_obj = datetime.strptime(target_date, '%Y-%m-%d')
        self.ws.cell(self.DATE_ROW, new_col, date_obj)
        self.logger.info(f"创建新日期列: {target_date} (列{new_col})")
        return new_col

    def update_data(self, code: str, date: str,
                   market_value: float, unit_price: float):
        """更新指定ETF的数据"""
        col_idx = self.find_or_create_date_column(date)

        # 只更新数据section
        for section in self.sections.values():
            if not section.is_data_section:
                continue

            row_idx = self._find_etf_row(code, section)
            if row_idx is None:
                self.logger.warning(
                    f"在section '{section.name}' 中未找到 {code}"
                )
                continue

            # 根据section类型更新对应的值
            value = market_value if '总市值' in section.name else unit_price
            self.ws.cell(row_idx, col_idx, value)
            self.logger.debug(
                f"更新 {code} {section.name}: {value}"
            )

    def recalculate_formulas(self, date: str):
        """
        重新计算所有公式单元格的值

        这个方法会读取"总市值"和"基金单位市值"的原始数据，
        然后计算其他sections的公式值并更新缓存
        """
        col_idx = self.find_or_create_date_column(date)

        # 获取所有ETF的总市值和单位市值数据
        etf_data = {}  # {code: {'market_value': float, 'unit_price': float, 'name': str}}

        # 从"总市值" section读取数据
        market_value_section = None
        for section in self.sections.values():
            if '总市值' in section.name:
                market_value_section = section
                break

        if market_value_section:
            for row in range(market_value_section.data_start, market_value_section.data_end + 1):
                code = self.ws.cell(row, self.CODE_COL).value
                name = self.ws.cell(row, self.NAME_COL).value
                market_value = self.ws.cell(row, col_idx).value

                if code and market_value is not None:
                    etf_data[code] = {
                        'market_value': float(market_value),
                        'name': name
                    }

        # 从"基金单位市值" section读取单位净值
        for section in self.sections.values():
            if '基金单位市值' in section.name:
                for row in range(section.data_start, section.data_end + 1):
                    code = self.ws.cell(row, self.CODE_COL).value
                    unit_price = self.ws.cell(row, col_idx).value

                    if code and code in etf_data and unit_price is not None:
                        etf_data[code]['unit_price'] = float(unit_price)
                break

        # 计算并更新其他sections
        for code, data in etf_data.items():
            if 'unit_price' not in data:
                continue

            market_value = data['market_value']
            unit_price = data['unit_price']

            # 计算基金份额（亿份）= 总市值（亿元）/ 单位净值（元）
            fund_share = market_value / unit_price if unit_price != 0 else 0

            # 更新各个计算型section
            for section in self.sections.values():
                if section.is_data_section:
                    continue  # 跳过原始数据section

                row_idx = self._find_etf_row(code, section)
                if row_idx is None:
                    continue

                # 根据section类型计算值
                if '基金份额' in section.name and '变动' not in section.name:
                    # 基金份额
                    value = fund_share
                elif '基金份额变动' in section.name and '比例' not in section.name:
                    # 基金份额变动：需要前一天的数据
                    prev_col = col_idx - 1
                    if prev_col >= self.DATA_START_COL:
                        prev_share_cell = self.ws.cell(row_idx, prev_col).value
                        if prev_share_cell is not None:
                            value = fund_share - float(prev_share_cell)
                        else:
                            value = None
                    else:
                        value = None
                elif '申赎净额' in section.name:
                    # 申赎净额：与份额变动相同
                    prev_col = col_idx - 1
                    if prev_col >= self.DATA_START_COL:
                        prev_share_cell = self.ws.cell(row_idx, prev_col).value
                        if prev_share_cell is not None:
                            value = fund_share - float(prev_share_cell)
                        else:
                            value = None
                    else:
                        value = None
                elif '份额变动比例' in section.name:
                    # 份额变动比例
                    prev_col = col_idx - 1
                    if prev_col >= self.DATA_START_COL:
                        prev_share_cell = self.ws.cell(row_idx, prev_col).value
                        if prev_share_cell is not None and float(prev_share_cell) != 0:
                            share_change = fund_share - float(prev_share_cell)
                            value = (share_change / float(prev_share_cell)) * 100
                        else:
                            value = None
                    else:
                        value = None
                elif '市值变动' in section.name:
                    # 市值变动
                    prev_col = col_idx - 1
                    if prev_col >= self.DATA_START_COL:
                        prev_mv_cell = self.ws.cell(row_idx, prev_col).value
                        if prev_mv_cell is not None:
                            value = market_value - float(prev_mv_cell)
                        else:
                            value = None
                    else:
                        value = None
                elif '涨跌幅' in section.name:
                    # 涨跌幅
                    prev_col = col_idx - 1
                    if prev_col >= self.DATA_START_COL:
                        prev_price_cell = self.ws.cell(row_idx, prev_col).value
                        if prev_price_cell is not None and float(prev_price_cell) != 0:
                            price_change = unit_price - float(prev_price_cell)
                            value = (price_change / float(prev_price_cell)) * 100
                        else:
                            value = None
                    else:
                        value = None
                else:
                    continue

                # 更新单元格值
                if value is not None:
                    self.ws.cell(row_idx, col_idx, value)
                    self.logger.debug(
                        f"计算 {code} {section.name}: {value}"
                    )

        self.logger.info(f"已重新计算日期 {date} 的所有公式值")

    def _find_etf_row(self, code: str, section: Section) -> Optional[int]:
        """在指定section中查找ETF行"""
        for row in range(section.data_start, section.data_end + 1):
            if self.ws.cell(row, self.CODE_COL).value == code:
                return row
        return None

    def save(self):
        """保存Excel文件"""
        # 设置标志，让Excel在打开时重新计算所有公式
        # 这样可以确保公式的缓存值是最新的
        self.wb.properties.calcMode = 'auto'

        # 标记工作簿需要重新计算
        if hasattr(self.ws, 'sheet_properties'):
            self.ws.sheet_properties.enableFormatConditionsCalculation = True

        self.wb.save(self.file_path)
        self.logger.info(f"Excel文件已保存: {self.file_path}")
