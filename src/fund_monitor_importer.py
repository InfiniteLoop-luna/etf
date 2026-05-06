from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


SHEET_NAME = "公募&私募基金"
CATEGORY_NAME_MAP = {
    "股票基金": "其中：股票基金",
    "混合基金": "其中：混合基金",
    "债券基金": "其中：债券基金",
    "货币基金": "其中：货币基金",
    "QDII基金": "其中：QDII基金",
    "公募证券投资基金合计": "合计",
    "权益类银行理财产品": "权益类理财产品",
}
SECTION_FIELD_MAP = {
    "当月情况": {
        "基金数量（只）": "fund_count",
        "份额（亿份）": "share_amount",
        "净值（亿元）": "nav_amount",
        "单位净值（元）": "unit_nav",
    },
    "环比变动情况": {
        "基金数量（只）": "mom_fund_count",
        "份额（亿份）": "mom_share_amount",
        "净值（亿元）": "mom_nav_amount",
        "单位净值（元）": "mom_unit_nav",
    },
    "同比变动情况": {
        "基金数量（只）": "yoy_fund_count",
        "份额（亿份）": "yoy_share_amount",
        "净值（亿元）": "yoy_nav_amount",
        "单位净值（元）": "yoy_unit_nav",
    },
}


def _normalize_text(value) -> str:
    if value is None:
        return ""
    return str(value).replace("\n", "").strip()


def _coerce_optional_float(value):
    if value is None:
        return None
    text = _normalize_text(value)
    if text in {"", "#DIV/0!", "#N/A"}:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _normalize_month_value(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return pd.to_datetime(value, unit="D", origin="1899-12-30").strftime("%Y-%m-%d")
    return pd.to_datetime(value).strftime("%Y-%m-%d")


def _normalize_category_name(value) -> str:
    category_name = _normalize_text(value)
    return CATEGORY_NAME_MAP.get(category_name, category_name)


def parse_fund_monitor_workbook(path) -> pd.DataFrame:
    workbook_source = Path(path) if isinstance(path, (str, Path)) else path
    wb = load_workbook(workbook_source, data_only=True, read_only=True)
    try:
        if SHEET_NAME not in wb.sheetnames:
            raise ValueError("导入文件缺少 公募&私募基金 sheet")

        ws = wb[SHEET_NAME]
        section_by_col = {}
        current_section = ""
        for col_idx in range(1, ws.max_column + 1):
            section = _normalize_text(ws.cell(2, col_idx).value)
            if section in SECTION_FIELD_MAP:
                current_section = section
            section_by_col[col_idx] = current_section

        field_by_col = {}
        for col_idx in range(1, ws.max_column + 1):
            field_by_col[col_idx] = _normalize_text(ws.cell(3, col_idx).value)

        rows = []
        current_month_text = None
        for row_idx in range(4, ws.max_row + 1):
            month_value = ws.cell(row_idx, 2).value
            if month_value is not None:
                current_month_text = _normalize_month_value(month_value)

            category_name = _normalize_category_name(ws.cell(row_idx, 3).value)
            if not category_name:
                continue
            if current_month_text is None:
                continue

            row = {"month": current_month_text, "category_name": category_name}
            for col_idx in range(4, ws.max_column + 1):
                section = section_by_col.get(col_idx, "")
                field = field_by_col.get(col_idx, "")
                mapped = SECTION_FIELD_MAP.get(section, {}).get(field)
                if not mapped:
                    continue
                row[mapped] = _coerce_optional_float(ws.cell(row_idx, col_idx).value)

            rows.append(row)

        if not rows:
            raise ValueError("未解析出任何基金监测记录")
        return pd.DataFrame(rows).sort_values(["month", "category_name"]).reset_index(drop=True)
    finally:
        wb.close()
