"""Excel数据加载器 - 将Excel数据转换为长格式DataFrame"""

import openpyxl
import pandas as pd
from datetime import datetime
from typing import Dict, List, Tuple
import logging


# 常量定义 - 与excel_manager.py保持一致
CODE_COL = 1
NAME_COL = 2
DATA_START_COL = 3
GLOBAL_DATE_ROW = 2  # 全局日期行，所有section共享同一个日期行


def load_etf_data(file_path: str) -> pd.DataFrame:
    """
    加载Excel文件并转换为长格式DataFrame

    Args:
        file_path: Excel文件路径

    Returns:
        DataFrame with columns: code, name, date, metric_type, value, is_aggregate
    """
    logger = logging.getLogger(__name__)
    logger.info(f"开始加载Excel文件: {file_path}")

    # 加载工作簿
    # 使用data_only=False以便读取公式，然后手动评估
    # 这样可以处理没有缓存值的公式单元格
    wb = openpyxl.load_workbook(file_path, data_only=False)
    ws = wb.active

    # 检测所有sections
    sections = _detect_sections(ws)
    logger.info(f"检测到 {len(sections)} 个sections")

    # 解析所有sections
    all_data = []
    for section_name, section_info in sections.items():
        logger.info(f"解析section: {section_name}")
        section_data = _parse_section(ws, section_name, section_info)
        logger.info(f"  -> 该section加载了 {len(section_data)} 条原始数据")
        all_data.extend(section_data)

    # 转换为DataFrame
    df = pd.DataFrame(all_data, columns=['code', 'name', 'date', 'metric_type', 'value', 'is_aggregate'])

    # 数据清洗
    logger.info(f"原始数据行数: {len(df)}")

    # 显示每个metric_type的原始数据量
    if len(df) > 0:
        logger.info("各指标原始数据量:")
        for metric in df['metric_type'].unique():
            count = len(df[df['metric_type'] == metric])
            logger.info(f"  {metric}: {count} 行")

    # 删除缺失值
    df = df.dropna()
    logger.info(f"删除缺失值后: {len(df)}")

    # 转换日期为datetime
    # 先检查日期格式
    if len(df) > 0:
        sample_dates = df['date'].head(10).tolist()
        logger.debug(f"示例日期值: {sample_dates}")

    df['date'] = pd.to_datetime(df['date'], errors='coerce')
    df = df.dropna(subset=['date'])
    logger.info(f"转换日期后: {len(df)}")

    # 转换值为数值类型
    df['value'] = pd.to_numeric(df['value'], errors='coerce')
    df = df.dropna(subset=['value'])
    logger.info(f"转换数值后: {len(df)}")

    logger.info(f"最终数据行数: {len(df)}")
    return df


def _detect_sections(ws) -> Dict[str, Dict]:
    """
    检测Excel中的所有sections

    Args:
        ws: openpyxl worksheet对象

    Returns:
        Dict[section_name, section_info]
        section_info包含: header_row, data_start, data_end
        注意：所有section共享全局的GLOBAL_DATE_ROW
    """
    sections = {}
    keywords = ['市值', '份额', '变动', '申赎', '比例', '涨跌幅']

    # 特殊处理：第一个section（总市值）没有header，直接从第3行开始
    # 检查第3行是否有ETF代码，如果有，说明存在这个特殊section
    first_data_row = 3
    if ws.cell(first_data_row, CODE_COL).value:
        # 找到第一个section的结束行（遇到空行或下一个section header）
        data_end = ws.max_row
        for row_idx in range(first_data_row, ws.max_row + 1):
            code_cell = ws.cell(row_idx, CODE_COL).value
            name_cell = ws.cell(row_idx, NAME_COL).value

            # 遇到空行或section header，说明第一个section结束
            if (code_cell is None and name_cell is None) or \
               (code_cell is None and name_cell and isinstance(name_cell, str) and
                any(kw in name_cell for kw in keywords)):
                data_end = row_idx - 1
                break

        sections['总市值'] = {
            'header_row': None,  # 没有header行
            'data_start': first_data_row,
            'data_end': data_end
        }

    # 检测其他有header的sections
    for row_idx in range(1, ws.max_row + 1):
        code_cell = ws.cell(row_idx, CODE_COL).value
        name_cell = ws.cell(row_idx, NAME_COL).value

        # Section header特征：CODE_COL为空，NAME_COL包含关键词
        if (code_cell is None and
            name_cell and
            isinstance(name_cell, str) and
            any(kw in name_cell for kw in keywords)):

            section_name = name_cell
            header_row = row_idx
            # 数据从header的下一行开始
            data_start = row_idx + 1

            # 找到数据结束行（下一个section开始或文件结束）
            data_end = ws.max_row
            for next_row in range(data_start, ws.max_row + 1):
                next_code = ws.cell(next_row, CODE_COL).value
                next_name = ws.cell(next_row, NAME_COL).value

                # 检查是否是下一个section的header
                if (next_code is None and
                    next_name and
                    isinstance(next_name, str) and
                    any(kw in next_name for kw in keywords)):
                    data_end = next_row - 1
                    break

            sections[section_name] = {
                'header_row': header_row,
                'data_start': data_start,
                'data_end': data_end
            }

    return sections


def _evaluate_simple_formula(ws, formula: str, row: int, col: int):
    """
    评估简单的Excel公式（主要是SUM公式和单元格引用）

    Args:
        ws: worksheet对象
        formula: 公式字符串（如 "=SUM(C3:C15)*22000/16000" 或 "=D34-C34"）
        row: 当前单元格行号
        col: 当前单元格列号

    Returns:
        计算结果，如果无法计算则返回None
    """
    import re

    try:
        # 移除开头的等号
        formula = formula.lstrip('=')

        # 处理单元格引用: A1, B2, etc.
        cell_pattern = r'([A-Z]+)(\d+)'

        def replace_cell_ref(match):
            col_letter = match.group(1)
            row_num = int(match.group(2))

            # 获取单元格值
            col_idx = openpyxl.utils.column_index_from_string(col_letter)
            cell = ws.cell(row_num, col_idx)
            cell_value = cell.value

            # 如果单元格本身是公式，递归评估
            if cell.data_type == 'f' and isinstance(cell_value, str):
                cell_value = _evaluate_simple_formula(ws, cell_value, row_num, col_idx)

            # 返回值
            if cell_value is None:
                return '0'
            elif isinstance(cell_value, (int, float)):
                return str(cell_value)
            else:
                return '0'

        # 先处理SUM函数: SUM(C3:C15)
        sum_pattern = r'SUM\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)'

        def replace_sum(match):
            start_col = match.group(1)
            start_row = int(match.group(2))
            end_col = match.group(3)
            end_row = int(match.group(4))

            # 计算列号
            start_col_idx = openpyxl.utils.column_index_from_string(start_col)
            end_col_idx = openpyxl.utils.column_index_from_string(end_col)

            # 求和
            total = 0
            for r in range(start_row, end_row + 1):
                for c in range(start_col_idx, end_col_idx + 1):
                    cell_value = ws.cell(r, c).value
                    if cell_value is not None and isinstance(cell_value, (int, float)):
                        total += cell_value

            return str(total)

        # 替换SUM函数
        formula = re.sub(sum_pattern, replace_sum, formula, flags=re.IGNORECASE)

        # 替换单元格引用
        formula = re.sub(cell_pattern, replace_cell_ref, formula)

        # 评估数学表达式
        result = eval(formula)
        return result

    except Exception:
        # 如果评估失败，返回None
        return None


def _parse_section(ws, section_name: str, section_info: Dict) -> List[Tuple]:
    """
    解析单个section的数据

    Args:
        ws: openpyxl worksheet对象
        section_name: section名称（作为metric_type）
        section_info: section信息（header_row, data_start, data_end）

    Returns:
        List of tuples: (code, name, date, metric_type, value, is_aggregate)
    """
    logger = logging.getLogger(__name__)
    data = []
    data_start = section_info['data_start']
    data_end = section_info['data_end']

    # 从全局日期行读取日期列表（从第3列开始）
    dates = []
    col_idx = DATA_START_COL
    while col_idx <= ws.max_column:
        date_val = ws.cell(GLOBAL_DATE_ROW, col_idx).value
        if date_val is None:
            break

        # 处理datetime对象
        if isinstance(date_val, datetime):
            dates.append(date_val.strftime("%Y-%m-%d"))
        # 处理字符串
        elif isinstance(date_val, str):
            # 标准化日期格式 (处理 2026/02/02 -> 2026-02-02)
            normalized = date_val.replace('/', '-')
            dates.append(normalized)
        else:
            # 其他类型尝试转换为字符串
            dates.append(str(date_val))

        col_idx += 1

    logger.debug(f"Section '{section_name}' 检测到 {len(dates)} 个日期列")
    if len(dates) > 0:
        logger.debug(f"前5个日期: {dates[:5]}")

    # 读取数据行
    for row_idx in range(data_start, data_end + 1):
        code = ws.cell(row_idx, CODE_COL).value
        name = ws.cell(row_idx, NAME_COL).value

        # 跳过空行
        if code is None and name is None:
            continue

        # 检查是否为汇总行
        is_aggregate = False
        if name and isinstance(name, str):
            aggregate_keywords = ['总计', '合计', '总和']
            if any(kw in name for kw in aggregate_keywords):
                is_aggregate = True
                code = 'ALL'  # 汇总行使用特殊代码

        # 读取该行的所有数据值
        for col_offset, date_str in enumerate(dates):
            col_idx = DATA_START_COL + col_offset
            cell = ws.cell(row_idx, col_idx)
            value = cell.value

            # 如果是公式单元格，尝试评估公式
            if cell.data_type == 'f' and value and isinstance(value, str):
                value = _evaluate_simple_formula(ws, value, row_idx, col_idx)

            # 跳过空值
            if value is None:
                continue

            # 添加数据记录
            data.append((
                code,
                name,
                date_str,
                section_name,
                value,
                is_aggregate
            ))

    return data
