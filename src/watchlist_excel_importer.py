from __future__ import annotations

import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from src.user_watchlist_store import add_watchlist_item, normalize_username


HEADER_ALIASES = {
    "code": {"代码", "证券代码", "股票代码", "ts_code", "TS_CODE", "code", "Code"},
    "name": {"名称", "证券名称", "股票名称", "简称", "name", "Name"},
}


def _normalize_cell_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).replace("\n", "").strip()


def normalize_stock_ts_code(raw_value) -> str:
    value = _normalize_cell_text(raw_value).upper()
    if not value:
        return ""

    value = value.replace(" ", "")
    if value.endswith(".0") and value[:-2].isdigit():
        value = value[:-2]

    prefix_match = re.fullmatch(r"(SH|SZ|BJ)(\d{6})", value)
    if prefix_match:
        return f"{prefix_match.group(2)}.{prefix_match.group(1)}"

    suffix_match = re.fullmatch(r"(\d{6})[.\-_]?(SH|SZ|BJ)", value)
    if suffix_match:
        return f"{suffix_match.group(1)}.{suffix_match.group(2)}"

    if value.isdigit() and len(value) <= 6:
        code = value.zfill(6)
        if code.startswith("6"):
            return f"{code}.SH"
        if code.startswith(("4", "8", "9")):
            return f"{code}.BJ"
        return f"{code}.SZ"

    return ""


def _find_header_columns(ws) -> tuple[int, int, int]:
    for row_idx in range(1, min(ws.max_row, 10) + 1):
        code_col = 0
        name_col = 0
        for col_idx in range(1, ws.max_column + 1):
            text = _normalize_cell_text(ws.cell(row_idx, col_idx).value)
            if text in HEADER_ALIASES["code"]:
                code_col = col_idx
            elif text in HEADER_ALIASES["name"]:
                name_col = col_idx
        if code_col:
            return row_idx, code_col, name_col or min(code_col + 1, ws.max_column)
    return 1, 1, 2 if ws.max_column >= 2 else 1


def parse_watchlist_import_workbook(source) -> list[dict]:
    workbook_source = Path(source) if isinstance(source, (str, Path)) else source
    if hasattr(workbook_source, "seek"):
        workbook_source.seek(0)

    wb = load_workbook(workbook_source, data_only=True, read_only=True)
    try:
        ws = wb.active
        header_row, code_col, name_col = _find_header_columns(ws)
        rows: list[dict] = []
        seen_codes: set[str] = set()

        for row_idx in range(header_row + 1, ws.max_row + 1):
            ts_code = normalize_stock_ts_code(ws.cell(row_idx, code_col).value)
            if not ts_code or ts_code in seen_codes:
                continue

            security_name = _normalize_cell_text(ws.cell(row_idx, name_col).value) if name_col else ""
            rows.append(
                {
                    "ts_code": ts_code,
                    "security_name": security_name or ts_code,
                    "security_type": "stock",
                    "source_row": row_idx,
                }
            )
            seen_codes.add(ts_code)

        if not rows:
            raise ValueError("No valid watchlist rows found in workbook")
        return rows
    finally:
        wb.close()


def import_watchlist_rows(
    username: str,
    rows: list[dict],
    *,
    existing_watchlist_df: pd.DataFrame | None = None,
    add_item=add_watchlist_item,
) -> dict:
    normalized_username = normalize_username(username)
    if not normalized_username:
        raise ValueError("username cannot be empty")

    existing_codes: set[str] = set()
    if (
        isinstance(existing_watchlist_df, pd.DataFrame)
        and not existing_watchlist_df.empty
        and "ts_code" in existing_watchlist_df.columns
    ):
        existing_codes = {
            str(code or "").strip().upper()
            for code in existing_watchlist_df["ts_code"].tolist()
            if str(code or "").strip()
        }

    added_codes: list[str] = []
    skipped_existing = 0
    skipped_invalid = 0
    failed_items: list[str] = []

    for row in rows or []:
        ts_code = normalize_stock_ts_code((row or {}).get("ts_code") or (row or {}).get("代码"))
        security_name = _normalize_cell_text(
            (row or {}).get("security_name")
            or (row or {}).get("name")
            or (row or {}).get("名称")
            or ts_code
        )
        security_type = _normalize_cell_text((row or {}).get("security_type") or "stock").lower() or "stock"

        if not ts_code:
            skipped_invalid += 1
            continue
        if ts_code in existing_codes:
            skipped_existing += 1
            continue

        try:
            add_item(
                normalized_username,
                ts_code,
                security_name=security_name or ts_code,
                security_type=security_type,
            )
        except Exception as exc:
            failed_items.append(f"{ts_code}: {exc}")
            continue

        existing_codes.add(ts_code)
        added_codes.append(ts_code)

    return {
        "added": len(added_codes),
        "added_codes": added_codes,
        "skipped_existing": skipped_existing,
        "skipped_invalid": skipped_invalid,
        "failed": len(failed_items),
        "failed_items": failed_items,
    }
