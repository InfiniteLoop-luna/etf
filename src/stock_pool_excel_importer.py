from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from src.user_stock_pool_store import format_tags
from src.watchlist_excel_importer import normalize_stock_ts_code


HEADER_ALIASES = {
    "code": {"代码", "证券代码", "股票代码", "ts_code", "TS_CODE", "code", "Code"},
    "name": {"名称", "证券名称", "股票名称", "简称", "name", "Name"},
    "industry": {"行业", "所属行业", "板块", "行业板块", "industry", "Industry", "sector"},
    "tags": {"标签", "分组", "自定义标签", "tag", "tags", "Tags", "group", "Group"},
    "note": {"备注", "说明", "理由", "note", "notes", "Note"},
}


def _normalize_cell_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).replace("\n", " ").strip()


def _find_header_columns(ws) -> tuple[int, dict[str, int]]:
    for row_idx in range(1, min(ws.max_row, 10) + 1):
        columns: dict[str, int] = {}
        for col_idx in range(1, ws.max_column + 1):
            text = _normalize_cell_text(ws.cell(row_idx, col_idx).value)
            for field, aliases in HEADER_ALIASES.items():
                if text in aliases and field not in columns:
                    columns[field] = col_idx
                    break
        if "code" in columns:
            if "name" not in columns and ws.max_column >= columns["code"] + 1:
                columns["name"] = columns["code"] + 1
            return row_idx, columns

    columns = {"code": 1}
    if ws.max_column >= 2:
        columns["name"] = 2
    return 1, columns


def parse_stock_pool_import_workbook(source) -> list[dict]:
    workbook_source = Path(source) if isinstance(source, (str, Path)) else source
    if hasattr(workbook_source, "seek"):
        workbook_source.seek(0)

    wb = load_workbook(workbook_source, data_only=True, read_only=True)
    try:
        ws = wb.active
        header_row, columns = _find_header_columns(ws)
        rows: list[dict] = []
        seen_codes: set[str] = set()

        for row_idx in range(header_row + 1, ws.max_row + 1):
            ts_code = normalize_stock_ts_code(ws.cell(row_idx, columns["code"]).value)
            if not ts_code or ts_code in seen_codes:
                continue

            security_name = (
                _normalize_cell_text(ws.cell(row_idx, columns["name"]).value)
                if columns.get("name")
                else ""
            )
            industry = (
                _normalize_cell_text(ws.cell(row_idx, columns["industry"]).value)
                if columns.get("industry")
                else ""
            )
            tags = (
                format_tags(_normalize_cell_text(ws.cell(row_idx, columns["tags"]).value))
                if columns.get("tags")
                else ""
            )
            note = (
                _normalize_cell_text(ws.cell(row_idx, columns["note"]).value)
                if columns.get("note")
                else ""
            )
            rows.append(
                {
                    "ts_code": ts_code,
                    "security_name": security_name or ts_code,
                    "industry": industry,
                    "tags": tags,
                    "note": note,
                    "source_row": row_idx,
                }
            )
            seen_codes.add(ts_code)

        if not rows:
            raise ValueError("未在 Excel 中找到有效股票代码")
        return rows
    finally:
        wb.close()
