from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


SHEET_NAME = "股票指数"
SECTION_FIELD_MAP = {
    "当月情况": {
        "涨幅": "monthly_change_pct",
        "开盘价格": "open_price",
        "收盘价格": "close_price",
        "最低点": "low_price",
        "最高点": "high_price",
        "期末静态市盈率": "static_pe",
        "期末动态市盈率": "dynamic_pe",
    },
    "环比变动情况": {
        "涨幅": "mom_change_pct",
        "开盘价格": "mom_open_price",
        "收盘价格": "mom_close_price",
        "最低点": "mom_low_price",
        "最高点": "mom_high_price",
        "期末静态市盈率": "mom_static_pe",
        "期末动态市盈率": "mom_dynamic_pe",
        "静态市盈率变化率": "mom_static_pe_change_rate",
        "动态市盈率变化率": "mom_dynamic_pe_change_rate",
    },
    "同比变动情况": {
        "涨幅": "yoy_change_pct",
        "开盘价格": "yoy_open_price",
        "收盘价格": "yoy_close_price",
        "最低点": "yoy_low_price",
        "最高点": "yoy_high_price",
        "期末静态市盈率": "yoy_static_pe",
        "期末动态市盈率": "yoy_dynamic_pe",
    },
}


def _normalize_header(value) -> str:
    if value is None:
        return ""
    return str(value).replace("\n", "").strip()


def parse_index_monitor_workbook(path) -> pd.DataFrame:
    workbook_source = Path(path) if isinstance(path, (str, Path)) else path
    wb = load_workbook(workbook_source, data_only=True, read_only=True)
    try:
        if SHEET_NAME not in wb.sheetnames:
            raise ValueError("导入文件缺少 股票指数 sheet")

        ws = wb[SHEET_NAME]
        row_iter = ws.iter_rows(values_only=True)
        _ = next(row_iter, ())
        section_row = next(row_iter, ())
        field_row = next(row_iter, ())

        max_len = max(len(section_row), len(field_row))
        section_by_col: dict[int, str] = {}
        current_section = ""
        for col_idx in range(1, max_len + 1):
            raw_section = section_row[col_idx - 1] if col_idx - 1 < len(section_row) else None
            section = _normalize_header(raw_section)
            if section in SECTION_FIELD_MAP:
                current_section = section
            section_by_col[col_idx] = current_section

        field_by_col: dict[int, str] = {}
        for col_idx in range(1, max_len + 1):
            raw_field = field_row[col_idx - 1] if col_idx - 1 < len(field_row) else None
            field_by_col[col_idx] = _normalize_header(raw_field)

        rows = []
        current_month_text = None
        for raw_row in row_iter:
            month_value = raw_row[1] if len(raw_row) > 1 else None
            if month_value is not None:
                current_month_text = pd.to_datetime(month_value).strftime("%Y-%m-%d")

            index_name = _normalize_header(raw_row[2] if len(raw_row) > 2 else None)
            if not index_name or index_name == "指数名称":
                continue
            if current_month_text is None:
                raise ValueError("股票指数 sheet 缺少月份")

            row = {"month": current_month_text, "index_name": index_name}
            for col_idx in range(4, max_len + 1):
                section = section_by_col.get(col_idx, "")
                field = field_by_col.get(col_idx, "")
                mapped_field = SECTION_FIELD_MAP.get(section, {}).get(field)
                if not mapped_field:
                    continue
                value = raw_row[col_idx - 1] if col_idx - 1 < len(raw_row) else None
                row[mapped_field] = float(value) if value is not None else None
            rows.append(row)

        if not rows:
            raise ValueError("未解析出任何指数记录")
        return pd.DataFrame(rows).sort_values(["month", "index_name"]).reset_index(drop=True)
    finally:
        wb.close()
