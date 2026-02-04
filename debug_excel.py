import openpyxl
import logging

logging.basicConfig(level=logging.DEBUG)

# 常量
CODE_COL = 1
NAME_COL = 2
DATA_START_COL = 3

print("正在打开Excel文件...")
wb = openpyxl.load_workbook('主要ETF基金份额变动情况.xlsx', data_only=False)
ws = wb.active

print(f"\nExcel文件信息:")
print(f"  总行数: {ws.max_row}")
print(f"  总列数: {ws.max_column}")

print(f"\n前20行的CODE_COL和NAME_COL内容:")
for row_idx in range(1, min(21, ws.max_row + 1)):
    code = ws.cell(row_idx, CODE_COL).value
    name = ws.cell(row_idx, NAME_COL).value
    print(f"  行{row_idx}: CODE={code}, NAME={name}")

print(f"\n检测sections...")
keywords = ['市值', '份额', '变动', '申赎', '比例', '涨跌幅']
sections_found = []

for row_idx in range(1, ws.max_row + 1):
    code_cell = ws.cell(row_idx, CODE_COL).value
    name_cell = ws.cell(row_idx, NAME_COL).value

    # Section header特征：CODE_COL为空，NAME_COL包含关键词
    if (code_cell is None and
        name_cell and
        isinstance(name_cell, str) and
        any(kw in name_cell for kw in keywords)):

        print(f"\n找到section: 行{row_idx}, 名称='{name_cell}'")
        sections_found.append((row_idx, name_cell))

        # 检查日期行
        date_row = row_idx + 1
        print(f"  日期行(行{date_row})前5列:")
        for col in range(DATA_START_COL, min(DATA_START_COL + 5, ws.max_column + 1)):
            date_val = ws.cell(date_row, col).value
            print(f"    列{col}: {date_val} (类型: {type(date_val).__name__})")

        # 检查数据行
        data_start = row_idx + 2
        print(f"  数据行(行{data_start}):")
        code = ws.cell(data_start, CODE_COL).value
        name = ws.cell(data_start, NAME_COL).value
        val1 = ws.cell(data_start, DATA_START_COL).value
        print(f"    CODE={code}, NAME={name}, 第一个值={val1} (类型: {type(val1).__name__})")

print(f"\n总共找到 {len(sections_found)} 个sections")
