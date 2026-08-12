import tempfile
import unittest
from datetime import datetime
from pathlib import Path

import openpyxl

from scripts.repair_etf_excel_gaps import EXPECTED_RAW_VALUES, find_incomplete_dates, raw_value_count
from src.data_loader import _evaluate_simple_formula


class EtfExcelGapRepairTest(unittest.TestCase):
    def test_formula_with_missing_source_is_not_zero(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = None
        ws['A2'] = 2
        self.assertIsNone(_evaluate_simple_formula(ws, '=A1/A2', 1, 2))

    def test_formula_with_complete_sources_still_calculates(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 6
        ws['A2'] = 2
        self.assertEqual(_evaluate_simple_formula(ws, '=A1/A2', 1, 2), 3)

    def test_finds_only_incomplete_raw_date(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(2, 3, datetime(2026, 8, 10))
        ws.cell(2, 4, datetime(2026, 8, 11))
        raw_rows = list(range(3, 16)) + list(range(19, 32))
        for row in raw_rows:
            ws.cell(row, 3, 1.0)
        ws.cell(3, 4, 1.0)
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / 'test.xlsx'
            wb.save(path)
            self.assertEqual(raw_value_count(ws, 3), EXPECTED_RAW_VALUES)
            self.assertEqual(
                find_incomplete_dates(path, datetime(2026, 8, 10).date(), datetime(2026, 8, 11).date()),
                ['2026-08-11'],
            )


if __name__ == '__main__':
    unittest.main()
