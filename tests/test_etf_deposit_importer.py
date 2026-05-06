import tempfile
import unittest
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook

from src.etf_deposit_importer import parse_deposit_workbook


class EtfDepositImporterTests(unittest.TestCase):
    def test_parse_deposit_workbook_reads_known_month_block(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "deposit.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "本外币存款数据"
            ws["B4"] = "2026-03-01"
            ws["C4"] = "人民币存款余额"
            ws["D4"] = 342.41
            ws["C5"] = "外币存款余额"
            ws["D5"] = 1.13
            ws["C6"] = "本外币存款余额"
            ws["D6"] = 350.23
            ws["C7"] = "住户存款增加额"
            ws["D7"] = 7.68
            ws["C8"] = "非金融企业存款增加额"
            ws["D8"] = 2.68
            ws["C9"] = "财政性存款增加额"
            ws["D9"] = 0.4606
            ws["C10"] = "非银行业金融机构存款增加额"
            ws["D10"] = 2.03
            ws["C11"] = "存款合计增加额"
            ws["D11"] = 13.73
            ws["C12"] = "居民长期贷款增加额"
            ws["D12"] = 0.4607
            wb.save(path)

            df = parse_deposit_workbook(path)

        self.assertEqual(df.iloc[0]["month"], "2026-03-01")
        self.assertAlmostEqual(df.iloc[0]["total_deposit_balance"], 350.23)

    def test_parse_deposit_workbook_raises_when_sheet_missing(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "wrong.xlsx"
            wb = Workbook()
            wb.active.title = "股票指数"
            wb.save(path)
            with self.assertRaisesRegex(ValueError, "本外币存款数据"):
                parse_deposit_workbook(path)

    def test_parse_deposit_workbook_accepts_file_like_object(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "本外币存款数据"
        ws["B4"] = "2026-03-01"
        ws["C4"] = "人民币存款余额"
        ws["D4"] = 342.41
        ws["C5"] = "外币存款余额"
        ws["D5"] = 1.13
        ws["C6"] = "本外币存款余额"
        ws["D6"] = 350.23
        ws["C7"] = "住户存款增加额"
        ws["D7"] = 7.68
        ws["C8"] = "非金融企业存款增加额"
        ws["D8"] = 2.68
        ws["C9"] = "财政性存款增加额"
        ws["D9"] = 0.4606
        ws["C10"] = "非银行业金融机构存款增加额"
        ws["D10"] = 2.03
        ws["C11"] = "存款合计增加额"
        ws["D11"] = 13.73
        ws["C12"] = "居民长期贷款增加额"
        ws["D12"] = 0.4607

        payload = BytesIO()
        wb.save(payload)
        payload.seek(0)

        df = parse_deposit_workbook(payload)

        self.assertEqual(df.iloc[0]["month"], "2026-03-01")


if __name__ == "__main__":
    unittest.main()
