import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from src.index_monitor_importer import parse_index_monitor_workbook


class IndexMonitorImporterTests(unittest.TestCase):
    def test_parse_index_monitor_workbook_reads_single_index_row(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "index-monitor.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "股票指数"
            ws["B4"] = "2026-05-01"
            ws["C3"] = "指数名称"
            ws["D2"] = "当月情况"
            ws["D3"] = "开盘价格"
            ws["E3"] = "收盘价格"
            ws["F3"] = "最低点"
            ws["G3"] = "最高点"
            ws["M3"] = "涨幅"
            ws["N2"] = "环比变动情况"
            ws["N3"] = "开盘价格"
            ws["O3"] = "收盘价格"
            ws["P3"] = "最低点"
            ws["Q3"] = "最高点"
            ws["W3"] = "涨幅"
            ws["X2"] = "同比变动情况"
            ws["X3"] = "开盘价格"
            ws["Y3"] = "收盘价格"
            ws["Z3"] = "最低点"
            ws["AA3"] = "最高点"
            ws["C4"] = "上证指数"
            ws["D4"] = 3340.12
            ws["E4"] = 3367.46
            ws["F4"] = 3321.18
            ws["G4"] = 3388.21
            ws["M4"] = 0.82
            ws["N4"] = 11.4
            ws["O4"] = 12.9
            ws["P4"] = -8.3
            ws["Q4"] = 20.1
            ws["W4"] = 5.13
            ws["X4"] = 188.4
            ws["Y4"] = 221.8
            ws["Z4"] = 160.5
            ws["AA4"] = 242.3
            wb.save(path)

            df = parse_index_monitor_workbook(path)

        self.assertEqual(df.iloc[0]["month"], "2026-05-01")
        self.assertEqual(df.iloc[0]["index_name"], "上证指数")
        self.assertAlmostEqual(df.iloc[0]["close_price"], 3367.46)
        self.assertAlmostEqual(df.iloc[0]["monthly_change_pct"], 0.82)
        self.assertAlmostEqual(df.iloc[0]["mom_change_pct"], 5.13)
        self.assertAlmostEqual(df.iloc[0]["yoy_close_price"], 221.8)

    def test_parse_index_monitor_workbook_raises_when_sheet_missing(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "wrong.xlsx"
            wb = Workbook()
            wb.active.title = "本外币存款数据"
            wb.save(path)
            with self.assertRaisesRegex(ValueError, "股票指数"):
                parse_index_monitor_workbook(path)


if __name__ == "__main__":
    unittest.main()
