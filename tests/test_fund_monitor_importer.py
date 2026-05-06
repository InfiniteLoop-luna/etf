import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from src.fund_monitor_importer import parse_fund_monitor_workbook


class FundMonitorImporterTests(unittest.TestCase):
    def test_parse_fund_monitor_workbook_reads_repeated_month_blocks(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "fund-monitor.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "公募&私募基金"
            ws["D2"] = "当月情况"
            ws["H2"] = "环比变动情况"
            ws["L2"] = "同比变动情况"
            headers = ["基金数量（只）", "份额（亿份）", "净值（亿元）", "单位净值（元）"]
            for idx, header in enumerate(headers, start=4):
                ws.cell(3, idx).value = header
                ws.cell(3, idx + 4).value = header
                ws.cell(3, idx + 8).value = header

            ws["B16"] = "2026-05-01"
            ws["C16"] = "其中：股票基金"
            ws["D16"] = 3585
            ws["E16"] = 39349.05
            ws["F16"] = 51128.57
            ws["G16"] = 1.2993
            ws["J16"] = -5168.49
            ws["N16"] = 6467.17

            ws["B22"] = "2026-04-01"
            ws["C22"] = "私募证券投资基金"
            ws["D22"] = 81745
            ws["F22"] = 74600
            ws["J22"] = 1100
            wb.save(path)

            df = parse_fund_monitor_workbook(path)

        self.assertEqual(
            df[["month", "category_name"]].values.tolist(),
            [
                ["2026-04-01", "私募证券投资基金"],
                ["2026-05-01", "其中：股票基金"],
            ],
        )
        self.assertAlmostEqual(df.iloc[1]["nav_amount"], 51128.57)
        self.assertAlmostEqual(df.iloc[0]["mom_nav_amount"], 1100)

    def test_parse_fund_monitor_workbook_normalizes_excel_month_serial_and_category_alias(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "fund-monitor-serial.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "公募&私募基金"
            ws["D2"] = "当月情况"
            ws["D3"] = "基金数量（只）"
            ws["E3"] = "份额（亿份）"
            ws["F3"] = "净值（亿元）"
            ws["G3"] = "单位净值（元）"
            ws["B5"] = 46113
            ws["C5"] = "股票基金"
            ws["F5"] = 51128.57
            wb.save(path)

            df = parse_fund_monitor_workbook(path)

        self.assertEqual(df.iloc[0]["month"], "2026-04-01")
        self.assertEqual(df.iloc[0]["category_name"], "其中：股票基金")

    def test_parse_fund_monitor_workbook_converts_excel_errors_to_none(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "fund-monitor-errors.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "公募&私募基金"
            ws["D2"] = "当月情况"
            ws["D3"] = "基金数量（只）"
            ws["E3"] = "份额（亿份）"
            ws["F3"] = "净值（亿元）"
            ws["G3"] = "单位净值（元）"
            ws["B16"] = "2026-05-01"
            ws["C16"] = "其中：货币基金"
            ws["G16"] = "#DIV/0!"
            wb.save(path)

            df = parse_fund_monitor_workbook(path)

        self.assertIsNone(df.iloc[0]["unit_nav"])

    def test_parse_fund_monitor_workbook_raises_when_sheet_missing(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "wrong.xlsx"
            wb = Workbook()
            wb.active.title = "股票指数"
            wb.save(path)
            with self.assertRaisesRegex(ValueError, "公募&私募基金"):
                parse_fund_monitor_workbook(path)


if __name__ == "__main__":
    unittest.main()
